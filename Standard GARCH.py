# -*- coding: utf-8 -*-
"""
Created on Mon Nov 23 13:53:12 2020

@author: Vaanis-Laptop
"""

import openpyxl

from arch import arch_model
import pandas as pd
pd.plotting.register_matplotlib_converters()
from matplotlib import pyplot
import seaborn as sns

new_wb = openpyxl.Workbook()
sheet = new_wb.active
 
path= "C:\\Users\\Vaanis-Laptop\\Documents\\Data.xlsx"
wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active 
max_r=sheet_obj.max_row

val=[]

for i in range(2, max_r):
    val.append((sheet_obj.cell(row = i,column= 5).value)*0.1)
    
returns = val
am = arch_model(returns,vol='Garch', p=1, o=1, q=1, dist='Normal')
res= am.fit(update_freq=5, disp='off')
new= res.forecast(horizon=10, start = 100)
new.variance[100:].plot()
print(new.variance.tail())

#print(res.summary())
